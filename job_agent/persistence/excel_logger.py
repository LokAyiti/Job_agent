"""Excel-based application log.

The log is the human-readable source of truth for which jobs were applied to.
It is also used by the Orchestrator to avoid duplicates.
"""
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from job_agent.models import JobApplication

COLUMNS = [
    "job_id",
    "title",
    "company",
    "location",
    "date_applied",
    "status",
    "link",
    "resume_path",
    "error_message",
    "notes",
    "retry_count",
    "fit_score",
    "source",
    "platform",
]


class ExcelLogger:
    def __init__(self, log_file: Path):
        self.log_file = log_file
        self.log_file.parent.mkdir(parents=True, exist_ok=True)
        if not self.log_file.exists():
            self._create_workbook()

    def _create_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(COLUMNS)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        wb.save(self.log_file)

    def _load(self):
        wb = load_workbook(self.log_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if headers != COLUMNS:
            # Migrate header row to the current schema.
            ws.delete_rows(1)
            ws.insert_rows(1)
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
            wb.save(self.log_file)
        return wb

    def list_applications(self) -> list[JobApplication]:
        wb = self._load()
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            row_dict = dict(zip(COLUMNS, row))
            try:
                rows.append(JobApplication.from_log_row(row_dict))
            except Exception:
                # Skip corrupted rows rather than crash the whole pipeline.
                continue
        return rows

    def get_application_by_id(self, job_id: str) -> JobApplication | None:
        for app in self.list_applications():
            if app.id == job_id:
                return app
        return None

    def upsert(self, application: JobApplication) -> None:
        wb = self._load()
        ws = wb.active
        row_data = application.to_log_row()
        values = [row_data[col] for col in COLUMNS]

        # Search for existing job_id.
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value == application.id:
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=idx, column=col_idx, value=value)
                break
        else:
            ws.append(values)

        wb.save(self.log_file)

    def upsert_many(self, applications: Iterable[JobApplication]) -> None:
        wb = self._load()
        ws = wb.active
        existing_ids = {}
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            existing_ids[row[0].value] = idx

        for app in applications:
            values = [app.to_log_row()[col] for col in COLUMNS]
            if app.id in existing_ids:
                idx = existing_ids[app.id]
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=idx, column=col_idx, value=value)
            else:
                ws.append(values)

        wb.save(self.log_file)

    def is_duplicate(self, application: JobApplication) -> bool:
        key = application.unique_key()
        for existing in self.list_applications():
            if existing.unique_key() == key and existing.status.value in {
                "submitted",
                "queued",
                "in_progress",
                "responded",
            }:
                return True
        return False
